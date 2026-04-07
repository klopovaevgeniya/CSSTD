from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Q
from core.decorators import admin_required
from core.models import Project, Employee, ProjectTask, Department, ProjectStatus
from io import BytesIO
import csv
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@admin_required
def reports_list(request):
    """Страница со списком отчетов"""
    return render(request, "admin/reports/list.html", {
        'excel_available': OPENPYXL_AVAILABLE,
    })


@admin_required
def export_projects_excel(request):
    """Экспорт проектов в Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Библиотека openpyxl не установлена", status=400)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Проекты"
    
    # Заголовки
    headers = ["ID", "Код", "Название", "Тип", "Статус", "Менеджер", "Дата начала", "Дата конца", "Бюджет", "Реальная стоимость"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    projects = Project.objects.select_related('manager', 'status', 'type')
    for row, project in enumerate(projects, 2):
        worksheet.cell(row=row, column=1).value = project.id
        worksheet.cell(row=row, column=2).value = project.code
        worksheet.cell(row=row, column=3).value = project.name
        worksheet.cell(row=row, column=4).value = project.type.name if project.type else ""
        worksheet.cell(row=row, column=5).value = project.status.name if project.status else ""
        manager_name = f"{project.manager.last_name} {project.manager.first_name}" if project.manager else ""
        worksheet.cell(row=row, column=6).value = manager_name
        worksheet.cell(row=row, column=7).value = project.start_date
        worksheet.cell(row=row, column=8).value = project.end_date
        worksheet.cell(row=row, column=9).value = float(project.budget) if project.budget else 0
        worksheet.cell(row=row, column=10).value = float(project.actual_cost) if project.actual_cost else 0
    
    # Ширину столбцов
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 18
    
    # Сохранение
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="projects_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response


@admin_required
def export_projects_csv(request):
    """Экспорт проектов в CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="projects_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(["ID", "Код", "Название", "Тип", "Статус", "Менеджер", "Дата начала", "Дата конца", "Бюджет", "Реальная стоимость"])
    
    projects = Project.objects.select_related('manager', 'status', 'type')
    for project in projects:
        manager_name = f"{project.manager.last_name} {project.manager.first_name}" if project.manager else ""
        writer.writerow([
            project.id,
            project.code,
            project.name,
            project.type.name if project.type else "",
            project.status.name if project.status else "",
            manager_name,
            project.start_date,
            project.end_date,
            float(project.budget) if project.budget else 0,
            float(project.actual_cost) if project.actual_cost else 0,
        ])
    
    return response


@admin_required
def export_employees_excel(request):
    """Экспорт сотрудников в Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Библиотека openpyxl не установлена", status=400)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Сотрудники"
    
    # Заголовки
    headers = ["ID", "ФИО", "Должность", "Отдел", "Email", "Телефон", "Дата найма", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    employees = Employee.objects.select_related('position', 'department')
    for row, employee in enumerate(employees, 2):
        fio = f"{employee.last_name} {employee.first_name}"
        if employee.middle_name:
            fio += f" {employee.middle_name}"
        
        worksheet.cell(row=row, column=1).value = employee.id
        worksheet.cell(row=row, column=2).value = fio
        worksheet.cell(row=row, column=3).value = employee.position.name if employee.position else ""
        worksheet.cell(row=row, column=4).value = employee.department.name if employee.department else ""
        worksheet.cell(row=row, column=5).value = employee.email
        worksheet.cell(row=row, column=6).value = employee.phone
        worksheet.cell(row=row, column=7).value = employee.hire_date
        worksheet.cell(row=row, column=8).value = "Активен" if employee.is_active else "Неактивен"
    
    # Ширина столбцов
    for i, width in enumerate([5, 30, 15, 15, 25, 15, 15, 12], 1):
        worksheet.column_dimensions[chr(64+i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response


@admin_required
def export_employees_csv(request):
    """Экспорт сотрудников в CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(["ID", "ФИО", "Должность", "Отдел", "Email", "Телефон", "Дата найма", "Статус"])
    
    employees = Employee.objects.select_related('position', 'department')
    for employee in employees:
        fio = f"{employee.last_name} {employee.first_name}"
        if employee.middle_name:
            fio += f" {employee.middle_name}"
        
        writer.writerow([
            employee.id,
            fio,
            employee.position.name if employee.position else "",
            employee.department.name if employee.department else "",
            employee.email,
            employee.phone,
            employee.hire_date,
            "Активен" if employee.is_active else "Неактивен",
        ])
    
    return response


@admin_required
def export_tasks_excel(request):
    """Экспорт задач в Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Библиотека openpyxl не установлена", status=400)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Задачи"
    
    # Заголовки
    headers = ["ID", "Проект", "Название", "Назначена", "Статус", "Приоритет", "Срок выполнения"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    tasks = ProjectTask.objects.select_related('project', 'assigned_to')
    for row, task in enumerate(tasks, 2):
        assigned_name = ""
        if task.assigned_to:
            assigned_name = f"{task.assigned_to.last_name} {task.assigned_to.first_name}"
        
        worksheet.cell(row=row, column=1).value = task.id
        worksheet.cell(row=row, column=2).value = task.project.name if task.project else ""
        worksheet.cell(row=row, column=3).value = task.name
        worksheet.cell(row=row, column=4).value = assigned_name
        worksheet.cell(row=row, column=5).value = task.status
        worksheet.cell(row=row, column=6).value = task.priority
        worksheet.cell(row=row, column=7).value = task.due_date
    
    # Ширина столбцов
    for i, width in enumerate([5, 25, 30, 20, 12, 12, 15], 1):
        worksheet.column_dimensions[chr(64+i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tasks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response


@admin_required
def export_tasks_csv(request):
    """Экспорт задач в CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="tasks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(["ID", "Проект", "Название", "Назначена", "Статус", "Приоритет", "Срок выполнения"])
    
    tasks = ProjectTask.objects.select_related('project', 'assigned_to')
    for task in tasks:
        assigned_name = ""
        if task.assigned_to:
            assigned_name = f"{task.assigned_to.last_name} {task.assigned_to.first_name}"
        
        writer.writerow([
            task.id,
            task.project.name if task.project else "",
            task.name,
            assigned_name,
            task.status,
            task.priority,
            task.due_date,
        ])
    
    return response
