from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Q, Sum, DecimalField, Value
from django.db.models.functions import Coalesce
from core.decorators import role_required
from core.models import Employee, Project, ProjectTask, ProjectParticipant
from datetime import date, timedelta
import csv
from io import BytesIO
from core.utils.project_archive import archived_project_q

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@role_required(['project_manager'])
def manager_statistics(request):
    """Страница со статистикой и отчётами менеджера"""
    employee = Employee.objects.filter(employee_user_id=request.session.get('user_id')).first()
    
    if not employee:
        projects = Project.objects.filter(manager__employee_user=request.user).exclude(archived_project_q())
        all_tasks = ProjectTask.objects.filter(project__manager__employee_user=request.user).exclude(
            archived_project_q(prefix='project')
        )
        team_members = ProjectParticipant.objects.filter(
            project__manager__employee_user=request.user
        ).exclude(
            archived_project_q(prefix='project')
        ).values('employee').distinct().count()
    else:
        projects = Project.objects.filter(manager=employee).exclude(archived_project_q()).select_related('status', 'type')
        all_tasks = ProjectTask.objects.filter(project__manager=employee).select_related(
            'status', 'project', 'assigned_to'
        ).exclude(
            archived_project_q(prefix='project')
        ).prefetch_related('task_assignees__employee')
        team_members = ProjectParticipant.objects.filter(
            project__manager=employee
        ).exclude(
            archived_project_q(prefix='project')
        ).values('employee').distinct().count()
    
    # Статистика по проектам
    total_projects = projects.count()
    active_projects = projects.exclude(status__name__icontains='завершён').count()
    completed_projects = projects.filter(status__name__icontains='завершён').count()
    
    # Статистика по задачам
    total_tasks = all_tasks.count()
    completed_tasks = all_tasks.filter(status__icontains='завершена').count()
    active_tasks = all_tasks.exclude(status__icontains='завершена').count()
    
    # Просроченные задачи
    overdue_tasks = all_tasks.filter(
        due_date__lt=date.today()
    ).exclude(status__icontains='завершена').count()
    
    # Задачи на сегодня
    today_tasks = all_tasks.filter(
        due_date=date.today()
    ).exclude(status__icontains='завершена').count()
    
    # Статистика по членам команды
    tasks_per_member = team_members / max(team_members, 1)
    
    # Бюджет
    total_budget = projects.aggregate(
        total=Coalesce(Sum('budget', output_field=DecimalField()), Value(0, output_field=DecimalField()))
    )['total'] or 0
    total_cost = projects.aggregate(
        total=Coalesce(Sum('actual_cost', output_field=DecimalField()), Value(0, output_field=DecimalField()))
    )['total'] or 0
    
    # Процент выполнения
    completion_percentage = int((completed_tasks / total_tasks * 100)) if total_tasks > 0 else 0
    
    # Задачи по приоритетам
    priority_stats = all_tasks.values('priority').annotate(count=Count('id')).order_by('-count')
    
    # Задачи по статусам
    status_stats = all_tasks.values('status').annotate(count=Count('id')).order_by('-count')
    
    # Проекты по типам
    project_type_stats = projects.values('type__name').annotate(count=Count('id')).order_by('-count')
    
    context = {
        'user': request.user,
        'employee': employee,
        'excel_available': OPENPYXL_AVAILABLE,
        
        # Статистика по проектам
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        
        # Статистика по задачам
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'active_tasks': active_tasks,
        'overdue_tasks': overdue_tasks,
        'today_tasks': today_tasks,
        'completion_percentage': completion_percentage,
        
        # Статистика по команде
        'team_members': team_members,
        'tasks_per_member': round(tasks_per_member, 1),
        
        # Бюджет
        'total_budget': f"{total_budget:,.2f}".replace(',', ' '),
        'total_cost': f"{total_cost:,.2f}".replace(',', ' '),
        
        # Детальная статистика
        'priority_stats': priority_stats,
        'status_stats': status_stats,
        'project_type_stats': project_type_stats,
    }
    
    return render(request, 'manager/statistics.html', context)


@role_required(['project_manager'])
def export_manager_projects_excel(request):
    """Экспорт проектов менеджера в Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Библиотека openpyxl не установлена", status=400)
    
    employee = Employee.objects.filter(employee_user_id=request.session.get('user_id')).first()
    
    if employee:
        projects = Project.objects.filter(manager=employee).exclude(archived_project_q()).select_related('status', 'type')
    else:
        projects = Project.objects.filter(manager__employee_user=request.user).exclude(archived_project_q()).select_related('status', 'type')
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Проекты"
    
    # Стиль для заголовка
    header_fill = PatternFill(start_color="264653", end_color="264653", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = ["ID", "Код", "Название", "Тип", "Статус", "Дата начала", "Дата конца", "Бюджет", "Реальная стоимость"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Данные
    for row, project in enumerate(projects, 2):
        worksheet.cell(row=row, column=1).value = project.id
        worksheet.cell(row=row, column=2).value = project.code or ""
        worksheet.cell(row=row, column=3).value = project.name
        worksheet.cell(row=row, column=4).value = project.type.name if project.type else ""
        worksheet.cell(row=row, column=5).value = project.status.name if project.status else ""
        worksheet.cell(row=row, column=6).value = project.start_date
        worksheet.cell(row=row, column=7).value = project.end_date
        worksheet.cell(row=row, column=8).value = float(project.budget) if project.budget else 0
        worksheet.cell(row=row, column=9).value = float(project.actual_cost) if project.actual_cost else 0
        
        # Применяем стили к ячейкам данных
        for col in range(1, 10):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            if col in [8, 9]:  # Числовые колонки
                cell.number_format = '#,##0.00'
    
    # Установка ширины колонок
    widths = [5, 12, 30, 15, 15, 15, 15, 15, 18]
    for col, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    
    # Сохранение
    from datetime import datetime
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="projects_manager_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response


@role_required(['project_manager'])
def export_manager_tasks_excel(request):
    """Экспорт задач менеджера в Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("Библиотека openpyxl не установлена", status=400)
    
    employee = Employee.objects.filter(employee_user_id=request.session.get('user_id')).first()
    
    if employee:
        tasks = ProjectTask.objects.filter(project__manager=employee).select_related(
            'project', 'assigned_to'
        ).exclude(
            archived_project_q(prefix='project')
        ).prefetch_related('task_assignees__employee')
    else:
        tasks = ProjectTask.objects.filter(project__manager__employee_user=request.user).select_related(
            'project', 'assigned_to'
        ).exclude(
            archived_project_q(prefix='project')
        ).prefetch_related('task_assignees__employee')
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Задачи"
    
    # Стиль для заголовка
    header_fill = PatternFill(start_color="e76f51", end_color="e76f51", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = ["ID", "Название", "Проект", "Статус", "Приоритет", "Назначена", "Срок выполнения", "Создано", "Обновлено"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Данные
    for row, task in enumerate(tasks, 2):
        worksheet.cell(row=row, column=1).value = task.id
        worksheet.cell(row=row, column=2).value = task.name
        worksheet.cell(row=row, column=3).value = task.project.name if task.project else ""
        worksheet.cell(row=row, column=4).value = task.status or ""
        worksheet.cell(row=row, column=5).value = task.priority or ""
        worksheet.cell(row=row, column=6).value = task.get_assignees_display()
        worksheet.cell(row=row, column=7).value = task.due_date
        worksheet.cell(row=row, column=8).value = task.created_at
        worksheet.cell(row=row, column=9).value = task.updated_at
        
        # Применяем стили к ячейкам данных
        for col in range(1, 10):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
    
    # Установка ширины колонок
    widths = [5, 25, 20, 12, 12, 20, 15, 15, 15]
    for col, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    
    # Сохранение
    from datetime import datetime
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tasks_manager_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response
